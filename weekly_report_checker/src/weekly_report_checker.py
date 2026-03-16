# -*- coding: utf-8 -*-
"""
Weekly Report Checker
讀取 OneNote 頁面 JSON（由 Get-OneNotePages.ps1 產出），
分析每位同仁的週報填寫狀況與內容，產出檢查結果 Excel。
"""

import argparse
import io
import json
import re
import sys
from collections import namedtuple
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from pathlib import Path

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        sys.stdout = io.TextIOWrapper(
            sys.stdout.buffer, encoding="utf-8", errors="replace"
        )

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    raise SystemExit("請先安裝依賴: pip install -r requirements.txt")

import sys as _sys
_sys.path.insert(0, str(Path(__file__).resolve().parent))
from onenote_parser import parse_page_xml, PageData, ReportItem, EXPECTED_HEADERS

DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent.parent / "output"
DEFAULT_WEEKS = 12

DATE_PATTERNS = [
    re.compile(r"^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$"),
    re.compile(r"^(\d{4})(\d{2})(\d{2})$"),
]

PageInfo = namedtuple("PageInfo", ["date", "content", "last_modified"])

FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_HEADER_BLUE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SECTION = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
FONT_RED = Font(color="9C0006")
FONT_GREEN = Font(color="006100")
FONT_ORANGE = Font(color="9C6500")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

HOLIDAYS_FILE = Path(__file__).resolve().parent.parent / "config" / "holidays.txt"


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def parse_date_from_page_name(name):
    if not name or not isinstance(name, str):
        return None
    s = name.strip()
    for pat in DATE_PATTERNS:
        m = pat.match(s)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return datetime(y, mo, d).date()
            except ValueError:
                continue
    return None


def parse_iso_datetime(s):
    """Parse ISO 8601 datetime string to datetime object (timezone-naive)."""
    if not s or not isinstance(s, str):
        return None
    s = s.replace("Z", "+00:00")
    for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z",
                "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.replace(tzinfo=None) if dt.tzinfo else dt
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def normalize_pages_list(pages):
    """Convert raw page list to list of PageInfo."""
    result = []
    if not isinstance(pages, list):
        return result
    for p in pages:
        if isinstance(p, str):
            d = parse_date_from_page_name(p)
            if d:
                result.append(PageInfo(date=d, content=None, last_modified=None))
        elif isinstance(p, dict):
            title = p.get("title") or p.get("displayName") or p.get("name") or ""
            d = parse_date_from_page_name(title)
            if not d:
                continue
            content = p.get("content") or p.get("body") or None
            lm_str = p.get("lastModifiedTime") or p.get("lastModified") or None
            last_modified = parse_iso_datetime(lm_str)
            result.append(PageInfo(date=d, content=content, last_modified=last_modified))
    return result


def get_member_pages_map(data):
    """
    從 JSON 取得 { 同仁名: [PageInfo, ...] }。
    支援格式：
    - { "members": [ { "name": "Paddy", "pages": [...] }, ... ] }
    - { "Paddy": ["2026/03/05", ...], ... }
    """
    out = {}
    if "members" in data and isinstance(data["members"], list):
        for m in data["members"]:
            if not isinstance(m, dict):
                continue
            name = m.get("同仁名") or m.get("member") or m.get("name") or ""
            if not name:
                continue
            pages = m.get("頁面日期") or m.get("pages") or m.get("pageDates") or []
            out[name] = normalize_pages_list(pages)
        return out
    for key, val in data.items():
        if key in ("members", "exportTime", "exportTimeUtc"):
            continue
        if isinstance(val, list):
            out[key] = normalize_pages_list(val)
        else:
            out[key] = []
    return out


def _get_member_contents_map(data):
    """Extract { member_name: [pageContent_dict, ...] } from JSON data."""
    out = {}
    if "members" in data and isinstance(data["members"], list):
        for m in data["members"]:
            if not isinstance(m, dict):
                continue
            name = m.get("name") or m.get("member") or m.get("同仁名") or ""
            if not name:
                continue
            contents = m.get("pageContents", [])
            if isinstance(contents, list) and contents:
                out[name] = contents
    return out


# ---------------------------------------------------------------------------
# Week helpers
# ---------------------------------------------------------------------------

def get_week_range(d):
    weekday = d.weekday()
    monday = d - timedelta(days=weekday)
    sunday = monday + timedelta(days=6)
    return monday, sunday


def get_this_week_and_last_week():
    today = datetime.now().date()
    this_monday, this_sunday = get_week_range(today)
    last_monday = this_monday - timedelta(days=7)
    last_sunday = last_monday + timedelta(days=6)
    return (this_monday, this_sunday), (last_monday, last_sunday)


def get_expected_report_dates(monday):
    """該週有效週報日期：週三、週四（同仁週三下班前 or 週四上班時填寫）。"""
    wed = monday + timedelta(days=2)
    thu = monday + timedelta(days=3)
    return {wed, thu}


def find_page_in_range(pages, start, end):
    """Return the first PageInfo whose date is Wed or Thu of the week (start=Monday, end=Sunday), or None."""
    valid = get_expected_report_dates(start)
    for p in pages:
        if p.date in valid:
            return p
    return None


# ---------------------------------------------------------------------------
# Analysis functions
# ---------------------------------------------------------------------------

def calc_consecutive_missing(pages, this_monday):
    """從本週往回數，連續幾週沒有對應頁面（有效日=該週三/四）。本週有填回傳 0。"""
    date_set = {p.date for p in pages}
    count = 0
    monday = this_monday
    while True:
        valid = get_expected_report_dates(monday)
        found = any(d in valid for d in date_set)
        if found:
            break
        count += 1
        monday -= timedelta(days=7)
        if count > 200:
            break
    return count


def calc_fill_rate(pages, this_monday, n_weeks=DEFAULT_WEEKS):
    """最近 n_weeks 週的填寫率（有效日=該週三/四），回傳 (filled_count, n_weeks)。"""
    date_set = {p.date for p in pages}
    filled = 0
    monday = this_monday
    for _ in range(n_weeks):
        valid = get_expected_report_dates(monday)
        if any(d in valid for d in date_set):
            filled += 1
        monday -= timedelta(days=7)
    return filled, n_weeks


def get_recent_missing_weeks(pages, this_monday, k=5):
    """回傳最近 k 週未填寫的週一日期清單（有效日=該週三/四，從本週往回算）。"""
    date_set = {p.date for p in pages}
    missing = []
    monday = this_monday
    for _ in range(k):
        valid = get_expected_report_dates(monday)
        if not any(d in valid for d in date_set):
            missing.append(monday)
        monday -= timedelta(days=7)
    return missing


def get_last_fill_date(pages):
    """回傳最近一筆頁面日期，或 None。"""
    if not pages:
        return None
    return max(p.date for p in pages)


def detect_late_submission(page):
    """
    檢查頁面是否遲交。
    遲交定義：lastModifiedTime 晚於該週週日 23:59。
    回傳遲了幾天，0 表示準時，None 表示無法判斷。
    """
    if page is None or page.last_modified is None:
        return None
    _, sunday = get_week_range(page.date)
    deadline = datetime.combine(sunday, datetime.max.time())
    if page.last_modified > deadline:
        delta = (page.last_modified.date() - sunday).days
        return max(delta, 1)
    return 0


def analyze_member(name, pages, this_range, last_range, n_weeks):
    """回傳 dict 包含所有分析指標。"""
    this_monday = this_range[0]
    this_page = find_page_in_range(pages, this_range[0], this_range[1])
    last_page = find_page_in_range(pages, last_range[0], last_range[1])

    # 指標 1：本週狀態
    late_days = detect_late_submission(this_page)
    if this_page is None:
        status = "未填寫"
    elif late_days and late_days > 0:
        status = "遲交（遲 {} 天）".format(late_days)
    else:
        status = "已填寫"

    # 指標 2：連續未填寫
    consecutive = calc_consecutive_missing(pages, this_monday)

    # 指標 3：填寫率
    filled, total = calc_fill_rate(pages, this_monday, n_weeks)
    rate = filled / total if total > 0 else 0

    # 指標 3-1：最近幾週未填寫的週期
    recent_missing_weeks = get_recent_missing_weeks(pages, this_monday, k=5)

    # 指標 4：最後填寫日期
    last_fill = get_last_fill_date(pages)

    # 備註
    notes = _build_notes(this_page, last_page, consecutive, late_days)

    return {
        "name": name,
        "this_week_status": status,
        "this_page_date": this_page.date if this_page else None,
        "consecutive_missing": consecutive,
        "fill_rate": rate,
        "filled_weeks": filled,
        "total_weeks": total,
        "last_fill_date": last_fill,
        "late_days": late_days,
        "recent_missing_weeks": recent_missing_weeks,
        "notes": notes,
    }


def _build_notes(this_page, last_page, consecutive, late_days):
    parts = []
    if this_page and last_page:
        parts.append("本週（{}）與上週（{}）皆有週報。".format(
            this_page.date.strftime("%Y/%m/%d"),
            last_page.date.strftime("%Y/%m/%d"),
        ))
    elif this_page:
        parts.append("本週有週報（{}），但上週未填寫。".format(
            this_page.date.strftime("%Y/%m/%d"),
        ))
    elif last_page:
        parts.append("本週未填寫（上週有：{}）。".format(
            last_page.date.strftime("%Y/%m/%d"),
        ))
    else:
        parts.append("本週與上週皆未填寫。")

    if consecutive >= 3:
        parts.append("已連續 {} 週未填寫！".format(consecutive))

    if late_days and late_days > 0:
        parts.append("本週週報遲交 {} 天。".format(late_days))

    return " ".join(parts)


# ---------------------------------------------------------------------------
# Holidays
# ---------------------------------------------------------------------------

def load_holidays(path=None):
    """Load holiday dates from config/holidays.txt (one date per line, YYYY/MM/DD)."""
    path = path or HOLIDAYS_FILE
    holidays = set()
    if not path.exists():
        return holidays
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            d = parse_date_from_page_name(line)
            if d:
                holidays.add(d)
    return holidays


# ---------------------------------------------------------------------------
# Content analysis: 9 rules + format check
# ---------------------------------------------------------------------------

def _fuzzy_match_items(items_a, items_b, threshold=0.6):
    """Match items between two weeks by fuzzy name similarity.
    Returns list of (item_a, item_b) pairs.  Unmatched items get None partner.
    """
    used_b = set()
    pairs = []
    for a in items_a:
        best_score = 0.0
        best_b = None
        best_idx = -1
        for idx, b in enumerate(items_b):
            if idx in used_b:
                continue
            score = SequenceMatcher(None, a.item, b.item).ratio()
            if score > best_score:
                best_score = score
                best_b = b
                best_idx = idx
        if best_score >= threshold and best_b is not None:
            pairs.append((a, best_b))
            used_b.add(best_idx)
        else:
            pairs.append((a, None))
    for idx, b in enumerate(items_b):
        if idx not in used_b:
            pairs.append((None, b))
    return pairs


def _detect_late_by_thursday(last_modified_str, page_date):
    """Rule 2: Check if update happened after Thursday 23:59 of the report week."""
    lm = parse_iso_datetime(last_modified_str)
    if lm is None or page_date is None:
        return None
    if isinstance(page_date, str):
        page_date = parse_date_from_page_name(page_date)
    if page_date is None:
        return None
    monday, _ = get_week_range(page_date)
    thursday_deadline = datetime.combine(monday + timedelta(days=3), datetime.max.time())
    if lm > thursday_deadline:
        delta = (lm.date() - (monday + timedelta(days=3))).days
        return max(delta, 1)
    return 0


def analyze_report_content(member_name, page_contents, holidays=None):
    """Analyze content of latest 2 pages for a member.
    page_contents: list of dicts with keys 'title', 'xml', 'lastModifiedTime'.
    Returns a dict with detailed analysis results.
    """
    holidays = holidays or set()
    result = {
        "member": member_name,
        "latest_date": None,
        "prev_date": None,
        "date_gap_days": None,
        "date_gap_ok": None,
        "skipped_week": False,
        "late_days": None,
        "format_ok": True,
        "format_issues": [],
        "item_checks": [],
        "rule_violations": {},
    }
    violations = {i: 0 for i in range(1, 10)}

    if not page_contents:
        return result

    pages_parsed = []
    for pc in page_contents[:2]:
        title = pc.get("title", "")
        xml_str = pc.get("xml", "")
        pd = parse_page_xml(xml_str, page_title=title)
        pages_parsed.append((title, pc, pd))

    latest_title, latest_pc, latest_pd = pages_parsed[0] if pages_parsed else ("", {}, PageData(title=""))
    prev_title, prev_pc, prev_pd = pages_parsed[1] if len(pages_parsed) > 1 else ("", {}, PageData(title=""))

    # Rule 1: dates
    latest_date = parse_date_from_page_name(latest_title)
    prev_date = parse_date_from_page_name(prev_title)
    result["latest_date"] = latest_date
    result["prev_date"] = prev_date

    if latest_date and prev_date:
        gap = (latest_date - prev_date).days
        result["date_gap_days"] = gap
        result["date_gap_ok"] = (gap == 7)
        if gap != 7:
            violations[1] += 1
        # Rule 3 & 4: skip week / holiday tolerance
        if gap > 7:
            result["skipped_week"] = True
            skip_monday = prev_date + timedelta(days=7)
            is_holiday_skip = False
            while skip_monday < latest_date:
                if skip_monday in holidays:
                    is_holiday_skip = True
                    break
                skip_monday += timedelta(days=7)
            if not is_holiday_skip:
                violations[3] += 1

    # Rule 2: Thursday deadline
    lm_str = latest_pc.get("lastModifiedTime", "")
    late_days = _detect_late_by_thursday(lm_str, latest_date)
    result["late_days"] = late_days
    if late_days and late_days > 0:
        violations[2] += 1

    # Format consistency check
    if latest_pd.headers and not latest_pd.header_match:
        result["format_ok"] = False
        result["format_issues"].append(
            "欄位不符: {} (預期 {})".format(latest_pd.headers, EXPECTED_HEADERS)
        )

    # Rule 5: leave tolerance (handled at the page-level analysis already)

    # Rules 6-9: Item-level comparison
    latest_items = latest_pd.items
    prev_items = prev_pd.items
    pairs = _fuzzy_match_items(latest_items, prev_items)

    for curr, prev in pairs:
        check = {
            "item_name": curr.item if curr else (prev.item if prev else "?"),
            "priority": curr.priority if curr else (prev.priority if prev else None),
            "prev_prgs": prev.progress_pct if prev else None,
            "curr_prgs": curr.progress_pct if curr else None,
            "prev_bw": prev.bw_spent_pct if prev else None,
            "curr_bw": curr.bw_spent_pct if curr else None,
            "has_weekly_note": bool(curr and curr.weekly_note),
            "issues": [],
        }

        if curr and prev:
            # Rule 6: progress check
            if curr.progress_pct is not None and prev.progress_pct is not None:
                if curr.progress_pct < prev.progress_pct:
                    check["issues"].append("進度下降 ({:.0f}%→{:.0f}%)".format(
                        prev.progress_pct * 100, curr.progress_pct * 100))
                    violations[6] += 1
            if curr.bw_spent_pct is not None and curr.bw_spent_pct > 0 and not curr.weekly_note:
                check["issues"].append("有工時但無週報內容")
                violations[6] += 1

            # Rule 7: BW > 0 but no progress and no note
            if (curr.bw_spent_pct is not None and curr.bw_spent_pct > 0
                    and curr.progress_pct is not None and prev.progress_pct is not None
                    and curr.progress_pct <= prev.progress_pct
                    and not curr.weekly_note):
                check["issues"].append("BW>0 但無進展且無報告")
                violations[7] += 1

            # Rule 8: 100% persists
            if (prev.progress_pct is not None and prev.progress_pct >= 1.0
                    and curr.progress_pct is not None and curr.progress_pct >= 1.0):
                check["issues"].append("已完成任務仍在週報中 (100%→100%)")
                violations[8] += 1

        if curr:
            # Rule 9: Priority 1 no effort
            if curr.priority == 1:
                no_bw = (curr.bw_spent_pct is None or curr.bw_spent_pct == 0)
                no_note = not curr.weekly_note
                if no_bw or no_note:
                    parts = []
                    if no_bw:
                        parts.append("BW=0%")
                    if no_note:
                        parts.append("無週報內容")
                    check["issues"].append("高優先任務無投入 ({})".format(", ".join(parts)))
                    violations[9] += 1

        if not curr and prev:
            check["issues"].append("上週有此 Item 但本週已移除")

        result["item_checks"].append(check)

    result["rule_violations"] = violations
    return result


# ---------------------------------------------------------------------------
# File finding
# ---------------------------------------------------------------------------

def find_latest_json(cwd, output_dir):
    candidates = []
    for folder in (cwd, output_dir):
        if not folder or not folder.exists():
            continue
        for f in folder.glob("*onenote_pages*.json"):
            try:
                candidates.append((f.stat().st_mtime, f))
            except OSError:
                pass
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(results, content_analyses, excel_path, n_weeks):
    wb = Workbook()
    _write_sheet1(wb, results, n_weeks)
    if content_analyses:
        _write_sheet2(wb, content_analyses)
    wb.save(excel_path)


def _write_sheet1(wb, results, n_weeks):
    ws = wb.active
    ws.title = "週報檢查"

    headers = [
        "同仁名",
        "本週狀態",
        "連續未填週數",
        "最近{}週填寫率".format(n_weeks),
        "最後填寫日期",
        "備註",
        "最近5週未填日期",
    ]
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = FONT_WHITE_BOLD
        cell.fill = FILL_HEADER_BLUE
        cell.alignment = header_align

    for row_idx, r in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=r["name"])

        status_cell = ws.cell(row=row_idx, column=2, value=r["this_week_status"])
        if r["this_week_status"] == "已填寫":
            status_cell.fill = FILL_GREEN
            status_cell.font = FONT_GREEN
        elif "遲交" in r["this_week_status"]:
            status_cell.fill = FILL_YELLOW
            status_cell.font = FONT_ORANGE
        else:
            status_cell.fill = FILL_RED
            status_cell.font = FONT_RED

        consec_cell = ws.cell(row=row_idx, column=3, value=r["consecutive_missing"])
        consec_cell.alignment = Alignment(horizontal="center")
        if r["consecutive_missing"] >= 3:
            consec_cell.fill = FILL_RED
            consec_cell.font = FONT_RED

        rate_str = "{:.0f}%（{}/{}）".format(
            r["fill_rate"] * 100, r["filled_weeks"], r["total_weeks"]
        )
        rate_cell = ws.cell(row=row_idx, column=4, value=rate_str)
        rate_cell.alignment = Alignment(horizontal="center")
        if r["fill_rate"] < 0.8:
            rate_cell.fill = FILL_RED
            rate_cell.font = FONT_RED
        elif r["fill_rate"] < 0.9:
            rate_cell.fill = FILL_YELLOW
            rate_cell.font = FONT_ORANGE

        last_fill_str = r["last_fill_date"].strftime("%Y/%m/%d") if r["last_fill_date"] else "—"
        ws.cell(row=row_idx, column=5, value=last_fill_str).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=r["notes"]).alignment = Alignment(wrap_text=True)

        # 最近 5 週未填寫日期（以週一日期表示）
        missing_weeks = r.get("recent_missing_weeks") or []
        if missing_weeks:
            missing_str = "、".join(d.strftime("%Y/%m/%d") for d in missing_weeks)
        else:
            missing_str = ""
        miss_cell = ws.cell(row=row_idx, column=7, value=missing_str)
        miss_cell.alignment = Alignment(wrap_text=True)
        if missing_weeks:
            miss_cell.fill = FILL_RED
            miss_cell.font = FONT_RED

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 30


def _write_sheet2(wb, content_analyses):
    """Write the '內容分析' sheet with per-member item-level checks."""
    ws = wb.create_sheet("內容分析")
    row = 1

    for ca in content_analyses:
        member = ca["member"]

        # --- Section header ---
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        sec_cell = ws.cell(row=row, column=1, value="【{}】".format(member))
        sec_cell.font = Font(bold=True, size=13)
        sec_cell.fill = FILL_SECTION
        row += 1

        # Meta info
        latest_str = ca["latest_date"].strftime("%Y/%m/%d") if ca["latest_date"] else "—"
        prev_str = ca["prev_date"].strftime("%Y/%m/%d") if ca["prev_date"] else "—"
        gap_str = "{}天".format(ca["date_gap_days"]) if ca["date_gap_days"] is not None else "—"

        ws.cell(row=row, column=1, value="最新頁面")
        ws.cell(row=row, column=2, value=latest_str)
        ws.cell(row=row, column=3, value="前一頁面")
        ws.cell(row=row, column=4, value=prev_str)
        ws.cell(row=row, column=5, value="間隔")
        ws.cell(row=row, column=6, value=gap_str)
        row += 1

        meta_parts = []
        if ca["skipped_week"]:
            meta_parts.append("跳週")
        if ca["late_days"] and ca["late_days"] > 0:
            meta_parts.append("遲交 {} 天（超過週四）".format(ca["late_days"]))
        if not ca["format_ok"]:
            meta_parts.append("格式不符: " + "; ".join(ca["format_issues"]))
        if meta_parts:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            note_cell = ws.cell(row=row, column=1, value="⚠ " + " | ".join(meta_parts))
            note_cell.font = FONT_ORANGE
            row += 1

        # Item table header
        item_headers = ["Priority", "Item", "上週Prgs", "本週Prgs", "BW%", "有週報內容", "檢查結果"]
        for col, h in enumerate(item_headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = FONT_WHITE_BOLD
            c.fill = FILL_HEADER_BLUE
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border = THIN_BORDER
        row += 1

        for chk in ca["item_checks"]:
            pri_str = str(chk["priority"]) if chk["priority"] is not None else ""
            prev_p = "{:.0f}%".format(chk["prev_prgs"] * 100) if chk["prev_prgs"] is not None else "—"
            curr_p = "{:.0f}%".format(chk["curr_prgs"] * 100) if chk["curr_prgs"] is not None else "—"
            bw_str = "{:.0f}%".format(chk["curr_bw"] * 100) if chk["curr_bw"] is not None else "—"
            note_yn = "✓" if chk["has_weekly_note"] else "✗"
            issues_str = "; ".join(chk["issues"]) if chk["issues"] else "OK"

            vals = [pri_str, chk["item_name"], prev_p, curr_p, bw_str, note_yn, issues_str]
            has_issue = bool(chk["issues"])
            for col, v in enumerate(vals, start=1):
                c = ws.cell(row=row, column=col, value=v)
                c.border = THIN_BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if has_issue:
                    c.fill = FILL_RED
                    c.font = FONT_RED
            row += 1

        # Summary
        violations = ca.get("rule_violations", {})
        total_violations = sum(violations.values())
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        summary_parts = []
        rule_names = {
            1: "日期間隔", 2: "週四前更新", 3: "跳週", 6: "進度/內容",
            7: "BW>0無產出", 8: "100%未移除", 9: "P1無投入",
        }
        for rid, count in sorted(violations.items()):
            if count > 0 and rid in rule_names:
                summary_parts.append("{}: {}項".format(rule_names[rid], count))
        if summary_parts:
            summary_text = "違規彙總: " + " | ".join(summary_parts)
        else:
            summary_text = "✓ 無異常"
        sum_cell = ws.cell(row=row, column=1, value=summary_text)
        if total_violations > 0:
            sum_cell.font = FONT_RED
        else:
            sum_cell.font = FONT_GREEN
        row += 2

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 45


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Weekly Report Checker：分析同仁週報填寫狀況，產出 Excel。"
    )
    parser.add_argument(
        "json_path",
        nargs="?",
        default=None,
        help="OneNote 頁面 JSON 檔案路徑；不填則自動尋找 output 下最新的 onenote_pages*.json",
    )
    parser.add_argument(
        "-o", "--output-dir",
        default=None,
        help="Excel 輸出目錄，預設為專案下的 output",
    )
    parser.add_argument(
        "-n", "--weeks",
        type=int,
        default=DEFAULT_WEEKS,
        help="填寫率計算的週數（預設 {}）".format(DEFAULT_WEEKS),
    )
    args = parser.parse_args()

    json_path = args.json_path
    if not json_path:
        cwd = Path.cwd()
        output_dir = Path(args.output_dir) if args.output_dir else DEFAULT_OUTPUT_DIR
        json_path = find_latest_json(cwd, output_dir)
        if not json_path:
            print("錯誤：未指定 JSON 檔案，且找不到 onenote_pages*.json。")
            return 1
        print("使用 JSON 檔案：", json_path)

    path = Path(json_path)
    if not path.exists():
        print("錯誤：找不到檔案", path)
        return 1

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    member_pages = get_member_pages_map(data)
    if not member_pages:
        print("錯誤：JSON 中未解析出任何同仁的頁面資料，請確認格式。")
        return 1

    this_range, last_range = get_this_week_and_last_week()
    n_weeks = args.weeks

    results = []
    for name, pages in member_pages.items():
        r = analyze_member(name, pages, this_range, last_range, n_weeks)
        results.append(r)

    # Content analysis (if pageContents available)
    holidays = load_holidays()
    content_analyses = []
    member_contents = _get_member_contents_map(data)
    for name in member_pages:
        pc_list = member_contents.get(name, [])
        if pc_list:
            ca = analyze_report_content(name, pc_list, holidays)
            content_analyses.append(ca)

    out_dir = Path(args.output_dir) if args.output_dir else DEFAULT_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = out_dir / "WeeklyReport_Check_{}.xlsx".format(ts)

    write_excel(results, content_analyses, excel_path, n_weeks)
    print("已產出：", excel_path)

    for r in results:
        print("  {}: {} | 連續未填:{}週 | 填寫率:{:.0f}%".format(
            r["name"], r["this_week_status"],
            r["consecutive_missing"], r["fill_rate"] * 100,
        ))

    if content_analyses:
        print("\n內容分析：")
        for ca in content_analyses:
            v = ca["rule_violations"]
            total = sum(v.values())
            print("  {}: 最新={} | 異常項={}".format(
                ca["member"],
                ca["latest_date"].strftime("%Y/%m/%d") if ca["latest_date"] else "—",
                total,
            ))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
